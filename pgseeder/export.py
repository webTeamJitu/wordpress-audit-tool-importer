from datetime import date

import psycopg
from config import *
from openpyxl import load_workbook
from psycopg.rows import dict_row

conninfo = f"user={DB_USER} password={DB_PASSWORD} host={DB_HOST} port={DB_PORT} dbname={DB_NAME} connect_timeout=10"


class wat_pg:
    def __init__(self) -> None:
        self.conn = psycopg.connect(conninfo)
        self.cur = self.conn.cursor(row_factory=dict_row)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.conn.closed:
            return

        if exc_type:
            try:
                self.conn.rollback()
            except Exception as exc2:
                print("error ignored in rollback on %s: %s", self, exc2)
        else:
            self.conn.commit()

        self.cur.close()
        self.conn.close()

    def insert(self, table, data):
        query = f"""
INSERT INTO public.{table} ({", ".join(data)})
    VALUES ({", ".join(map(lambda key: f"%({key})s", data))});
"""
        self.cur.execute(query, data)

    def exec_sp(self, sp, data={}):
        query = f"CALL {sp}({self.get_kwargs(data)})"
        self.cur.execute(query, data)

    def exec_fn(self, fn, data={}):
        query = f"SELECT * FROM {fn}({self.get_kwargs(data)})"
        self.cur.execute(query, data)
        return self.cur

    @staticmethod
    def get_kwargs(data):
        return ", ".join(map(lambda key: f"_{key}=>%({key})s", data))


class wat_xslx:
    def __init__(self, db) -> None:
        self.wb = load_workbook(filename=WORKBOOK_PATH, data_only=True)
        self.client_wss = {ws: ws["A35"].value for ws in self.wb.worksheets[:-1]}
        self.date_cells = next(
            self.wb.worksheets[0].iter_cols(min_col=DATE_COL, max_col=DATE_COL)
        )[3:]
        self.db = db

    def export_clients(self):
        for ws, url in self.client_wss.items():
            __client = {"username": "", "password": "", "logo_path": ""}
            __client = {"url": url, **__client}

            if _client := self.db.exec_fn("fn_ClientsGet", {"url": url}).fetchone():
                client = {"id": _client["fn_ClientsGet"]["guid"], "name": ws.title, **__client}
            else:
                client = {"name": ws.title, **__client}

            self.db.exec_sp("sp_ClientInsertOrUpdate", client)
        self.db.conn.commit()

    def export_login_retries(self):
        for ws, url in self.client_wss.items():
            _client = self.db.exec_fn("fn_ClientsGet", {"url": url}).fetchone()
            for cell in self.date_cells:
                retries = ws.cell(cell.row, RETRIES_COL)
                if retries.value is None:
                    continue

                self.db.insert(
                    "login_retries",
                    {
                        "retries": retries.value,
                        "date": cell.value,
                        "client_id": _client["fn_ClientsGet"]["id"],
                    },
                )
        self.db.conn.commit()

    def export_sites_availability(self):
        __availability = {"status_code": None, "reason": None}
        for ws, url in self.client_wss.items():
            _client = self.db.exec_fn("fn_ClientsGet", {"url": url}).fetchone()
            for cell in self.date_cells:
                latency = ws.cell(cell.row, LATENCY_COL).value
                downtime = ws.cell(cell.row, DOWNTIME_COL).value
                if latency is None and downtime is None:
                    continue

                self.db.insert(
                    "sites_availability",
                    {
                        **__availability,
                        "latency": latency or 0,
                        "downtime": downtime or 0,
                        "date_added": cell.value,
                        "client_id": _client["fn_ClientsGet"]["id"],
                    },
                )
        self.db.conn.commit()


with wat_pg() as db:
    xl = wat_xslx(db)
    xl.export_clients()
    xl.export_login_retries()
    xl.export_sites_availability()
