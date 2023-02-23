import psycopg
from psycopg.rows import dict_row
from openpyxl import load_workbook

import json
import datetime

DATE_COL = 1
RETRIES_COL = 2
LATENCY_COL = 6
DOWNTIME_COL = 11


def _get_kwargs(data):
    return ", ".join(map(lambda i: f"_{i}=>%({i})s", data))


def insert(cur, table, data):
    query = f"""
        INSERT INTO public.{table} ({", ".join(data)})
        VALUES ({", ".join(map(lambda v: f"%({v})s", data))});
        """
    cur.execute(query, data)


def fnExec(cur, fn, data: dict = {}):
    query = f"SELECT * FROM {fn}({_get_kwargs(data)})"
    cur.execute(query, data)


def spExec(cur, sp, data: dict = {}):
    query = f"CALL {sp}({_get_kwargs(data)})"
    cur.execute(query, data)


with open(".env.json") as f:
    env = json.load(f)

with open("sites.json") as f:
    tmp = f.read()

conninfo = "user={DB_USER} password={DB_PASSWORD} host={DB_HOST} port={DB_PORT} dbname={DB_NAME} connect_timeout=10"
conninfo = conninfo.format(**env)

wb = load_workbook(filename="./feb 2023 daily monitoring.xlsx", data_only=True)


def get_client(url):
    fnExec(cur, "fn_ClientsGet", {"url": url})
    return cur.fetchone()


def insert_or_update_clients():
    for ws in wb.worksheets[:-1]:
        client_name = ws.title
        client_url = ws["A35"].value

        temp_client = {**json.loads(tmp)["clientInfo"], "url": client_url}

        if _client := get_client(client_url):
            client = {**temp_client, "id": _client["id"], "name": client_name}
        else:
            client = {**temp_client, "name": client_name}

        spExec(cur, "sp_ClientInsertOrUpdate", client)
    conn.commit()


def insert_login_retries():
    for ws in wb.worksheets[:-1]:
        date_cells = next(ws.iter_cols(min_col=DATE_COL, max_col=DATE_COL))
        client_url = ws["A35"].value

        # temp_retries = json.loads(tmp)["loginRetries"]

        if _client := get_client(client_url):
            for cell in date_cells:
                if isinstance(cell.value, datetime.datetime):
                    retries = ws.cell(cell.row, RETRIES_COL)
                    if retries.value is None:
                        continue
                    retries = {
                        # **temp_retries,
                        "retries": retries.value,
                        "date": cell.value,
                        "client_id": _client["id"],
                    }
                    insert(cur, "login_retries", retries)
    conn.commit()


def insert_sites_availability():
    for ws in wb.worksheets[:-1]:
        date_cells = next(ws.iter_cols(min_col=DATE_COL, max_col=DATE_COL))
        client_url = ws["A35"].value

        temp_availability = json.loads(tmp)["siteAvailability"]

        if _client := get_client(client_url):
            for cell in date_cells:
                if isinstance(cell.value, datetime.datetime):
                    latency = ws.cell(cell.row, LATENCY_COL)
                    downtime = ws.cell(cell.row, DOWNTIME_COL)
                    if latency.value is None and downtime.value is None:
                        continue
                    availability = {
                        **temp_availability,
                        "latency": latency.value or 0,
                        "downtime": downtime.value or 0,
                        "date_added": cell.value,
                        "client_id": _client["id"],
                    }
                    insert(cur, "sites_availability", availability)
    conn.commit()


with psycopg.connect(conninfo) as conn:
    with conn.cursor(row_factory=dict_row) as cur:
        insert_or_update_clients()
        insert_login_retries()
        insert_sites_availability()
